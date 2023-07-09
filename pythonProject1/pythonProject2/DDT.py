import openpyxl

path = r"H:\New folder\Data.xlsx"
workbook = openpyxl.load_workbook(path)
sheet1 = workbook["Sheet1"]
rows = sheet1.max_row
colums = sheet1.max_column

# for r in range(2,rows+1):
#     for c in range(1,colums+1):
#         print(sheet1.cell(r,c).value, end= " ")
#     print(" ")

# Writing data into excel
path2 = r"H:\New folder\data2.xlsx"
sheet2 = workbook.active  # (or) sheet = workbook["Data"]

for r in range(1,6):
    for c in range(1,4):
        sheet2.cell(r,c).value = "welcome"
workbook